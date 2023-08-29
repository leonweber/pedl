import string
from collections import defaultdict
from operator import itemgetter
from typing import Optional, Dict, List, Set, Tuple
from argparse import ArgumentParser
from pathlib import Path
import random

from lxml import etree
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
import pandas as pd
import numpy as np
import requests
from openpyxl.worksheet.table import Table, TableStyleInfo
from omegaconf import DictConfig
from tqdm import tqdm, trange
import hydra
from Bio import Entrez

np.random.seed(42)
random.seed(42)

ESEARCH_MAX_COUNT = 100


def get_pmid_to_mesh_terms(pmids: Set[str], mesh_terms: Set[str]) -> Dict[str, List[str]]:
    """
    Get a mapping of PubMed IDs to their associated MeSH terms.

    This function retrieves MeSH terms for a given set of PubMed IDs (pmids) by querying the PubMed database.
    It then filters the MeSH terms based on a provided set of terms (mesh_terms) and returns a dictionary
    mapping each PubMed ID to a list of its associated MeSH terms.

    Args:
    pmids (Set[str]): A set of unique PubMed IDs for which MeSH terms are to be fetched.
    mesh_terms (Set[str]): A set of MeSH terms to filter the results by.

    Returns:
    Dict[str, List[str]]: A dictionary mapping each PubMed ID (key) to a list of its associated MeSH terms (value).

    Examples:
    >>> pmids = {"12345678", "23456789"}
    >>> mesh_terms = {"Disease", "Therapeutics"}
    >>> get_pmid_to_mesh_terms(pmids, mesh_terms)
    {'12345678': ['Disease'], '23456789': ['Therapeutics']}
    """
    pmid_to_mesh_terms = defaultdict(list)

    pmids_list = list(pmids)
    for i in trange(0, len(pmids_list), ESEARCH_MAX_COUNT, desc="Fetching MeSH terms"):
        batch_pmids = pmids_list[i:i + ESEARCH_MAX_COUNT]
        pmids_str = ",".join(batch_pmids)

        Entrez.email = "leonweber@cis.lmu.de"
        Entrez.tool = "PEDL"
        handle = Entrez.efetch(db="pubmed", id=pmids_str, retmode="xml")
        tree = etree.parse(handle)

        for article in tree.xpath("//PubmedArticle"):
            pmid = article.xpath("MedlineCitation/PMID")[0].text
            mesh_terms_pmid = set(article.xpath("MedlineCitation/MeshHeadingList/MeshHeading/DescriptorName/text()"))
            mesh_terms_pmid = mesh_terms_pmid.intersection(mesh_terms)  # TODO We might change this to also match subterms

            pmid_to_mesh_terms[pmid] = sorted(mesh_terms_pmid)

    return pmid_to_mesh_terms


def adjust_sheet_width(sheet):
    # Adjust column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 255)
        sheet.column_dimensions[column].width = adjusted_width


def _add_value_choices(sheet):
    # Incorrect
    sheet["P1"].value = "Wrong Genes"
    sheet["P2"].value = "Wrong PPA Type"
    sheet["P3"].value = "No PPA"

    data_validation = DataValidation(type="list", formula1="$P$1:$P$100")
    data_validation.add("K2:K10000")
    sheet.add_data_validation(data_validation)

    # Useless
    sheet["O1"].value = "Indirect Interaction"
    sheet["O2"].value = "Insufficient Evidence"
    sheet["O3"].value = "Wrong Organism"
    sheet["O4"].value = "Wrong Disease"
    sheet["O5"].value = "Mutation Required"
    sheet["O6"].value = "PTM Required"
    sheet["O7"].value = "Bound Proteins Required"

    data_validation = DataValidation(type="list", formula1="$O$1:$O$100")
    data_validation.add("L2:L10000")
    sheet.add_data_validation(data_validation)


def _add_table(sheet, num_rows, num_cols):
    last_col = string.ascii_uppercase[num_cols - 1]
    ref = f"A1:{last_col}{num_rows}"
    i = 1
    while True:
        try:
            table = Table(displayName=f"Table{i}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            sheet.add_table(table)
            break
        except ValueError:
            i += 1


def fill_sheet_from_df(
    sheet,
    df: pd.DataFrame,
    ppa_dir: Path,
    top_k_articles: int = 5,
    threshold: float = 0.5,
    pmid_to_mesh_terms: Optional[Dict] = None,
    annotation_mode: bool = False,
):
    header = [
        "head",
        "association type",
        "tail",
        "text",
        "pubmed",
        "article score",
        "total score",
        "mean score",
        "MESH terms",
    ]
    if annotation_mode:
        header += ["correct", "useful", "why incorrect?", "why not useful?", "comment"]
    for cell, value in zip(sheet["A1":"Z1"][0], header):
        cell.value = value
        cell.font = Font(bold=True)

    idx_next_free_row = 0

    all_ppa_data = []

    for i, (_, row) in enumerate(df.iterrows()):
        ppa_data = {"row": row, "score": 0}
        fname = ppa_dir / ((row["head"] + "-_-" + row["tail"]).upper() + ".txt")
        assert fname.exists(), f"{fname} does not exist"
        ppa_data["pmid_to_score"] = defaultdict(float)
        ppa_data["pmid_to_fields"] = defaultdict(list)

        with fname.open() as f:
            for line in f:
                fields = line.strip().split("\t")
                if len(fields) > 1:
                    ppa_type, score, pmid, text, _ = fields
                    if float(score) < threshold or ppa_type != row["association type"]:
                        continue

                    if pmid_to_mesh_terms and not pmid_to_mesh_terms[pmid]:
                        continue

                    ppa_data["pmid_to_score"][pmid] += float(score)
                    ppa_data["pmid_to_fields"][pmid].append(fields)
                    ppa_data["score"] += float(score)
        all_ppa_data.append(ppa_data)

    for ppa_data in sorted(all_ppa_data, key=itemgetter("score"), reverse=True):
        for i, (pmid, article_score) in enumerate(sorted(
            ppa_data["pmid_to_score"].items(), key=itemgetter(1), reverse=True
        )[:top_k_articles]):
            fields = ppa_data["pmid_to_fields"][pmid][0]
            ppa_type, score, pmid, text, _ = fields
            sheet[f"A{idx_next_free_row + 2}"].value = ppa_data["row"]["head"]
            sheet[f"C{idx_next_free_row + 2}"].value = ppa_data["row"]["tail"]
            sheet[f"B{idx_next_free_row + 2}"].value = ppa_data["row"][
                "association type"
            ]
            sheet[f"D{idx_next_free_row + 2}"].value = text
            sheet[
                f"E{idx_next_free_row + 2}"
            ].hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            sheet[f"E{idx_next_free_row + 2}"].value = pmid
            sheet[f"B{idx_next_free_row + 2}"].value = ppa_data["row"][
                "association type"
            ]
            sheet[f"E{idx_next_free_row + 2}"].style = "Hyperlink"
            sheet[f"F{idx_next_free_row + 2}"].value = article_score
            if i == 0:
                sheet[f"F{idx_next_free_row + 2}"].font = Font(bold=True)

            sheet[f"G{idx_next_free_row + 2}"].value = ppa_data["row"]["score (sum)"]

            sheet[f"H{idx_next_free_row + 2}"].value = ppa_data["row"]["score (sum)"] / ppa_data["row"]["num_found"]

            if pmid_to_mesh_terms and pmid in pmid_to_mesh_terms:
                sheet[f"I{idx_next_free_row + 2}"].value = ", ".join(
                    pmid_to_mesh_terms[pmid]
                )
            idx_next_free_row += 1

    if annotation_mode:
        _add_value_choices(sheet)
    adjust_sheet_width(sheet)
    if idx_next_free_row > 0:
        _add_table(sheet, num_rows=idx_next_free_row + 1, num_cols=len(header))


def _get_ppas_from_sheet(sheet):
    ppas = set()

    it = zip(sheet["A"], sheet["B"], sheet["C"])
    next(it) # skip header
    for head, rel, tail in it:
        ppa = (head.value, rel.value, tail.value)
        if all(ppa):
            ppas.add(ppa)

    return ppas



def _add_summary_sheet(df_summary, sheet, wb):
    ppas_in_sheet = _get_ppas_from_sheet(sheet)
    summary_sheet = wb.create_sheet(title=("Summary_" + sheet.title)[:31])
    n_rows_added = 0
    for i, row in enumerate(
            dataframe_to_rows(df_summary.sort_values("score (sum)", ascending=False),
                              index=False, header=True)):
        if i == 0 or tuple(row[:3]) in ppas_in_sheet:
            summary_sheet.append(row)
            n_rows_added += 1
    adjust_sheet_width(summary_sheet)
    _add_table(summary_sheet, num_rows=n_rows_added,
               num_cols=len(df_summary.columns))


def build_summary_table(
    raw_dir: Path, threshold: float = 0.0, no_association_type: bool = False
) -> pd.DataFrame:
    df = {
        "head": [],
        "association type": [],
        "tail": [],
        "score (sum)": [],
        "score (mean)": [],
        "score (max)": [],
        "pmids": [],
        "num_found": [],
    }


    rel_to_score_sum = defaultdict(float)
    rel_to_score_max = defaultdict(float)
    rel_to_num_found = defaultdict(int)
    rel_to_pmids = defaultdict(set)

    # hgnc_symbols = set(get_hgnc_symbol_to_gene_id().keys())

    files = list(raw_dir.glob("*.txt"))
    for file in tqdm(files):
        with file.open() as f:
            p1, p2 = file.name.replace(".txt", "").split("-_-")
            for line in f:
                fields = line.strip().split()
                if fields:
                    if no_association_type:
                        p1_unified, p2_unified = sorted([p1, p2])
                        rel = (p1_unified, "association", p2_unified)
                    else:
                        rel = (p1, fields[0], p2)
                    if float(fields[1]) >= threshold:
                        rel_to_score_sum[rel] += float(fields[1])
                        rel_to_num_found[rel] += 1
                        rel_to_score_max[rel] = max(
                            float(fields[1]), rel_to_score_max[rel]
                        )
                        rel_to_pmids[rel].add(fields[2])

    for rel, score_sum in rel_to_score_sum.items():
        score_max = rel_to_score_max[rel]
        pmids = ",".join(rel_to_pmids[rel])
        df["head"].append(rel[0])
        df["association type"].append(rel[1])
        df["tail"].append(rel[2])
        df["score (sum)"].append(score_sum)
        df["score (max)"].append(score_max)
        df["score (mean)"].append(score_sum / rel_to_num_found[rel])
        df["pmids"].append(pmids)
        df["num_found"].append(rel_to_num_found[rel])

    return pd.DataFrame(df)


def summarize_csv(cfg):
    df_summary = build_summary_table(raw_dir=Path(cfg.input), threshold=cfg.threshold,
                                     no_association_type=cfg.no_association_type)
    df_summary.to_csv(Path(cfg.output).with_suffix(".csv"), index=False)


def summarize_excel(cfg):
    df_summary = build_summary_table(raw_dir=Path(cfg.input), threshold=cfg.threshold)
    all_pmids = set()
    for pmids_rel in df_summary["pmids"]:
        all_pmids.update(pmids_rel.split(","))
    if cfg.mesh_terms:
        pmid_to_mesh_terms = get_pmid_to_mesh_terms(all_pmids, set(cfg.mesh_terms))
    else:
        pmid_to_mesh_terms = None

    wb = Workbook()
    sheet = wb.active

    if cfg.entity_set:
        with cfg.entity_set.open() as f:
            entity_set = set(f.read().split("\n"))
            df_a_to_b = df_summary[df_summary["head"].isin(entity_set)]
            df_b_to_a = df_summary[df_summary["tail"].isin(entity_set)]
            sheet.title = f"{cfg.entity_set.with_suffix('').name} -> other"[:31]
            fill_sheet_from_df(
                sheet=sheet,
                df=df_a_to_b,
                ppa_dir=Path(cfg.input),
                threshold=cfg.threshold,
                top_k_articles=cfg.top_k_articles,
                pmid_to_mesh_terms=pmid_to_mesh_terms,
            )
            _add_summary_sheet(df_a_to_b, sheet, wb)

            sheet = wb.create_sheet(title=f"other -> {cfg.entity_set.with_suffix('').name}"[:31])
            fill_sheet_from_df(
                sheet=sheet,
                df=df_b_to_a,
                ppa_dir=Path(cfg.input),
                threshold=cfg.threshold,
                top_k_articles=cfg.top_k_articles,
                pmid_to_mesh_terms=pmid_to_mesh_terms,
            )
            _add_summary_sheet(df_b_to_a, sheet, wb)
    else:
        fill_sheet_from_df(
            sheet=sheet,
            df=df_summary,
            ppa_dir=Path(cfg.input),
            threshold=cfg.threshold,
            top_k_articles=cfg.top_k_articles,
            pmid_to_mesh_terms=pmid_to_mesh_terms,
        )

        _add_summary_sheet(df_summary, sheet, wb)

    output_file = Path(cfg.output).with_suffix(".xlsx")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


@hydra.main(config_path="configs", config_name="summarize.yaml", version_base=None)
def summarize(cfg: DictConfig):
    if cfg.plain:
        summarize_csv(cfg)
    else:
        summarize_excel(cfg)



if __name__ == "__main__":
    summarize()