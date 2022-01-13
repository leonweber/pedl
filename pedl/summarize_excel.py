import string
from collections import defaultdict
from operator import itemgetter
from typing import Optional, Dict, List, Set
from argparse import ArgumentParser
from pathlib import Path
import random

from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
import pandas as pd
import numpy as np
import requests
from openpyxl.worksheet.table import Table, TableStyleInfo

from pedl.utils import build_summary_table

np.random.seed(42)
random.seed(42)

ESEARCH_MAX_COUNT = 100000


def get_pmid_to_mesh_terms(mesh_terms: Set[str]) -> Dict[str, List[str]]:
    pmid_to_mesh_terms = defaultdict(list)
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"

    for mesh_term in mesh_terms:
        n_processed = 0
        result = requests.get(
            base_url,
            params={
                "db": "pubmed",
                "term": f"{mesh_term}[MESH]",
                "tool": "PEDL",
                "email": "weberple@hu-berlin.de",
                "retmode": "json",
                "retmax": ESEARCH_MAX_COUNT,
            },
        )
        total_count = int(result.json()["esearchresult"]["count"])
        if total_count == 0:
            print(
                f"PubMed search for {mesh_term}[MESH] did not yield any articles. Is this really a valid mesh term?"
            )
        n_processed += len(result.json()["esearchresult"]["idlist"])
        for pmid in result.json()["esearchresult"]["idlist"]:
            pmid_to_mesh_terms[pmid].append(mesh_term)

        while n_processed < total_count:
            result = requests.get(
                base_url,
                params={
                    "db": "pubmed",
                    "term": f"{mesh_term}[MESH]",
                    "tool": "PEDL",
                    "email": "weberple@hu-berlin.de",
                    "retmode": "json",
                    "retmax": ESEARCH_MAX_COUNT,
                    "retstart": n_processed,
                },
            )
            n_processed += len(result.json()["esearchresult"]["idlist"])
            for pmid in result.json()["esearchresult"]["idlist"]:
                pmid_to_mesh_terms[pmid].append(mesh_term)

    return pmid_to_mesh_terms


def adjust_sheet_width(sheet):
    # Adjust column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 255)
        sheet.column_dimensions[column].width = adjusted_width


def _add_value_choices(sheet):
    # Incorrect
    sheet["P1"].value = "Wrong Genes"
    sheet["P2"].value = "Wrong PPA Type"
    sheet["P3"].value = "No PPA"

    data_validation = DataValidation(type="list", formula1="$P$1:$P$100")
    data_validation.add("K2:K10000")
    sheet.add_data_validation(data_validation)

    # Useless
    sheet["O1"].value = "Indirect Interaction"
    sheet["O2"].value = "Insufficient Evidence"
    sheet["O3"].value = "Wrong Organism"
    sheet["O4"].value = "Wrong Disease"
    sheet["O5"].value = "Mutation Required"
    sheet["O6"].value = "PTM Required"
    sheet["O7"].value = "Bound Proteins Required"

    data_validation = DataValidation(type="list", formula1="$O$1:$O$100")
    data_validation.add("L2:L10000")
    sheet.add_data_validation(data_validation)


def _add_table(sheet, num_rows, num_cols):
    last_col = string.ascii_uppercase[num_cols - 1]
    ref = f"A1:{last_col}{num_rows}"
    i = 1
    while True:
        try:
            table = Table(displayName=f"Table{i}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            sheet.add_table(table)
            break
        except ValueError:
            i += 1



def fill_sheet_from_df(
    sheet,
    df: pd.DataFrame,
    ppa_dir: Path,
    topk: int = 5,
    threshold: float = 0.5,
    pmid_to_mesh_terms: Optional[Dict] = None,
    annotation_mode: bool = False,
):
    header = [
        "head",
        "association type",
        "tail",
        "text",
        "pubmed",
        "article score",
        "PPA score (total)",
        "MESH terms",
    ]
    if annotation_mode:
        header += ["correct", "useful", "why incorrect?", "why not useful?", "comment"]
    for cell, value in zip(sheet["A1":"Z1"][0], header):
        cell.value = value
        cell.font = Font(bold=True)

    idx_next_free_row = 0

    all_ppa_data = []

    for i, (_, row) in enumerate(df.iterrows()):
        ppa_data = {"row": row, "score": 0}
        fname = ppa_dir / ((row["head"] + "-_-" + row["tail"]).upper() + ".txt")
        assert fname.exists(), f"{fname} does not exist"
        ppa_data["pmid_to_score"] = defaultdict(float)
        ppa_data["pmid_to_fields"] = defaultdict(list)

        with fname.open() as f:
            for line in f:
                fields = line.strip().split("\t")
                if len(fields) > 1:
                    ppa_type, score, pmid, text, _ = fields
                    if float(score) < threshold or ppa_type != row["association type"]:
                        continue

                    if pmid_to_mesh_terms and pmid not in pmid_to_mesh_terms:
                        continue

                    ppa_data["pmid_to_score"][pmid] += float(score)
                    ppa_data["pmid_to_fields"][pmid].append(fields)
                    ppa_data["score"] += float(score)
        all_ppa_data.append(ppa_data)

    for ppa_data in sorted(all_ppa_data, key=itemgetter("score"), reverse=True):
        for i, (pmid, article_score) in enumerate(sorted(
            ppa_data["pmid_to_score"].items(), key=itemgetter(1), reverse=True
        )[:topk]):
            fields = ppa_data["pmid_to_fields"][pmid][0]
            ppa_type, score, pmid, text, _ = fields
            sheet[f"A{idx_next_free_row + 2}"].value = ppa_data["row"]["head"]
            sheet[f"C{idx_next_free_row + 2}"].value = ppa_data["row"]["tail"]
            sheet[f"B{idx_next_free_row + 2}"].value = ppa_data["row"][
                "association type"
            ]
            sheet[f"D{idx_next_free_row + 2}"].value = text
            sheet[
                f"E{idx_next_free_row + 2}"
            ].hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            sheet[f"E{idx_next_free_row + 2}"].value = pmid
            sheet[f"B{idx_next_free_row + 2}"].value = ppa_data["row"][
                "association type"
            ]
            sheet[f"E{idx_next_free_row + 2}"].style = "Hyperlink"
            sheet[f"F{idx_next_free_row + 2}"].value = article_score
            if i == 0:
                sheet[f"F{idx_next_free_row + 2}"].font = Font(bold=True)

            sheet[f"G{idx_next_free_row + 2}"].value = ppa_data["row"]["score (sum)"]

            if pmid_to_mesh_terms and pmid in pmid_to_mesh_terms:
                sheet[f"H{idx_next_free_row + 2}"].value = ", ".join(
                    pmid_to_mesh_terms[pmid]
                )
            idx_next_free_row += 1

    if annotation_mode:
        _add_value_choices(sheet)
    adjust_sheet_width(sheet)
    if idx_next_free_row > 0:
        _add_table(sheet, num_rows=idx_next_free_row + 1, num_cols=len(header))


def summarize_excel(args):
    if args.mesh_terms:
        pmid_to_mesh_terms = get_pmid_to_mesh_terms(set(args.mesh_terms))
    else:
        pmid_to_mesh_terms = None

    df_summary = build_summary_table(raw_dir=args.ppa_dir, score_cutoff=args.threshold)

    wb = Workbook()
    sheet = wb.active

    if args.set_a:
        with args.set_a.open() as f:
            set_a = set(f.read().split("\n"))
            df_a_to_b = df_summary[df_summary["head"].isin(set_a)]
            df_b_to_a = df_summary[df_summary["tail"].isin(set_a)]
            sheet.title = f"{args.set_a.with_suffix('').name} -> other"[:31]
            fill_sheet_from_df(
                sheet=sheet,
                df=df_a_to_b,
                ppa_dir=args.ppa_dir,
                threshold=args.threshold,
                topk=args.topk,
                pmid_to_mesh_terms=pmid_to_mesh_terms,
                annotation_mode=args.annotation,
            )

            summary_sheet = wb.create_sheet(title=("Summary_" + sheet.title)[:31])
            for row in dataframe_to_rows(df_a_to_b.sort_values("score (sum)", ascending=False),
                                         index=False, header=True):
                summary_sheet.append(row)
            _add_table(summary_sheet, num_rows=len(df_a_to_b),
                       num_cols=len(df_a_to_b.columns))

            sheet = wb.create_sheet(title=f"other -> {args.set_a.with_suffix('').name}"[:31])
            fill_sheet_from_df(
                sheet=sheet,
                df=df_b_to_a,
                ppa_dir=args.ppa_dir,
                threshold=args.threshold,
                topk=args.topk,
                pmid_to_mesh_terms=pmid_to_mesh_terms,
                annotation_mode=args.annotation,
            )

            summary_sheet = wb.create_sheet(title=("Summary_" + sheet.title)[:31])
            for row in dataframe_to_rows(df_b_to_a.sort_values("score (sum)", ascending=False),
                                         index=False, header=True):
                summary_sheet.append(row)
            _add_table(summary_sheet, num_rows=len(df_b_to_a),
                       num_cols=len(df_b_to_a.columns))
    else:
        fill_sheet_from_df(
            sheet=sheet,
            df=df_summary,
            ppa_dir=args.ppa_dir,
            threshold=args.threshold,
            topk=args.topk,
            pmid_to_mesh_terms=pmid_to_mesh_terms,
            annotation_mode=args.annotation,
        )

        summary_sheet = wb.create_sheet(title="Summary")
        for row in dataframe_to_rows(df_summary.sort_values("score (sum)", ascending=False),
                                     index=False, header=True):
            summary_sheet.append(row)
        _add_table(summary_sheet, num_rows=len(df_summary),
                   num_cols=len(df_summary.columns))

    output_file = args.output.with_suffix(".xlsx")
    wb.save(output_file)
